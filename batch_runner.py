#!/usr/bin/env python
# -*- coding: utf-8 -*-

import subprocess
import itertools
import os
import time
import logging
import pandas as pd
from datetime import datetime
from tqdm import tqdm
import concurrent.futures
import threading
import json
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# 导入模型配置
from model_config import MODEL_CONFIG, TRY_TIMES_RANGE, DEFAULT_MAX_RETRIES, DEFAULT_MAX_WORKERS, SERVICE_NAME

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('batch_run.log'),
        logging.StreamHandler()
    ]
)

# 定义要遍历的参数
# 使用从配置文件导入的参数
model_config = MODEL_CONFIG
try_times_range = TRY_TIMES_RANGE

# 每个脚本的最大重试次数
MAX_RETRIES = DEFAULT_MAX_RETRIES

# 并行线程数
MAX_WORKERS = DEFAULT_MAX_WORKERS

# 结果表格文件名
RESULTS_EXCEL = 'ladder_match_results.xlsx'

# 结果记录（使用线程锁保护）
results = []
results_lock = threading.Lock()

# 已有结果（用于断点续测）
existing_results = {}

# 创建结果目录
results_dir = 'batch_results'
os.makedirs(results_dir, exist_ok=True)

# 当前时间作为运行ID
run_id = datetime.now().strftime('%Y%m%d_%H%M%S')
run_id = ""
run_dir = os.path.join(results_dir, run_id) if run_id else results_dir
os.makedirs(run_dir, exist_ok=True)

# 加载已有结果表格（用于断点续测）
def load_existing_results():
    if not os.path.exists(RESULTS_EXCEL):
        return {}
    
    results_data = {}
    try:
        wb = openpyxl.load_workbook(RESULTS_EXCEL)
        ws = wb.active
        
        # 读取表头
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value)
        
        # 读取每行数据
        for row in range(2, ws.max_row + 1):
            model_name = ws.cell(row=row, column=1).value
            if not model_name:
                continue
                
            results_data[model_name] = {}
            
            for col in range(2, len(headers) + 1):
                column_header = headers[col - 1]
                if column_header != "模型名称":  # 跳过第一列的模型名称
                    try:
                        # 提取提示类型和尝试次数
                        parts = column_header.split('_try')
                        prompt_type = parts[0]
                        try_times = int(parts[1]) if len(parts) > 1 else 1
                        
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value is not None:
                            # 保存格式为 {model_name: {prompt_type: {try_times: rank}}}
                            if prompt_type not in results_data[model_name]:
                                results_data[model_name][prompt_type] = {}
                            results_data[model_name][prompt_type][try_times] = cell_value
                    except Exception as e:
                        logging.warning(f"解析表格数据时出错: {e}")
        
        logging.info(f"已加载现有结果数据: {len(results_data)} 个模型")
        return results_data
    except Exception as e:
        logging.error(f"加载结果表格失败: {e}")
        return {}

# 更新结果表格
def update_results_excel(all_results):
    # 准备表头：模型名称 + 所有提示类型和尝试次数组合
    headers = ["模型名称"]
    prompt_try_combinations = []
    
    for model, prompts in model_config.items():
        for prompt in prompts:
            for try_times in try_times_range:
                column_name = f"{prompt}_try{try_times}"
                if column_name not in prompt_try_combinations:
                    prompt_try_combinations.append(column_name)
    
    # 按提示类型排序
    prompt_try_combinations.sort()
    headers.extend(prompt_try_combinations)
    
    # 创建数据框
    df_data = []
    
    # 找出最大层数，用于颜色范围计算
    max_rank = 1  # 默认至少是1，避免除以0
    for model_name, results in all_results.items():
        for prompt_type, try_results in results.items():
            for try_times, rank in try_results.items():
                if isinstance(rank, (int, float)) and rank > 0 and rank > max_rank:
                    max_rank = rank
    
    # logging.info(f"检测到的最大层数: {max_rank}")
    
    # 计算颜色范围
    def get_color_for_rank(rank):
        if rank == -1:  # 错误
            return "FF0000"  # 红色
        elif rank == 0:  # 失败
            return "FFC7CE"  # 浅红色
        elif rank > 0:  # 成功，颜色从浅绿色渐变到深绿色
            # 计算颜色范围从浅绿色(200,255,200)到深绿色(0,128,0)
            # 根据rank/max_rank的比例插值
            ratio = rank / max_rank
            r = int(200 - ratio * 200)  # 200->0
            g = int(255 - ratio * 127)  # 255->128
            b = int(200 - ratio * 200)  # 200->0
            return f"{r:02X}{g:02X}{b:02X}"
        return "FFFFFF"  # 默认白色
    
    for model_name, results in all_results.items():
        row = {"模型名称": model_name}
        
        # 填充每个提示类型和尝试次数的结果
        for prompt_type, try_results in results.items():
            for try_times, rank in try_results.items():
                column_name = f"{prompt_type}_try{try_times}"
                row[column_name] = rank
        
        df_data.append(row)
    
    # 创建DataFrame
    df = pd.DataFrame(df_data)
    
    # 如果数据框为空，添加表头
    if df.empty:
        df = pd.DataFrame(columns=headers)
    
    # 确保所有列都存在
    for col in headers:
        if col not in df.columns:
            df[col] = None
    
    # 按模型名称排序
    df = df.sort_values("模型名称")
    
    # 将DataFrame保存为Excel
    if os.path.exists(RESULTS_EXCEL):
        # 如果文件已存在，尝试保留单元格格式
        wb = openpyxl.load_workbook(RESULTS_EXCEL)
        ws = wb.active
        
        # 清空工作表
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
        
        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
        
        # 写入数据
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, header in enumerate(headers, 1):
                if header in row:
                    ws.cell(row=row_idx, column=col_idx, value=row[header])
                    
                    # 根据值设置颜色
                    if header != "模型名称" and row[header] is not None:
                        value = row[header]
                        color = get_color_for_rank(value)
                        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # 调整列宽
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) * 1.2, 12)
        
        # 保存文件
        wb.save(RESULTS_EXCEL)
    else:
        # 新建文件
        df.to_excel(RESULTS_EXCEL, index=False)
        
        # 添加格式
        wb = openpyxl.load_workbook(RESULTS_EXCEL)
        ws = wb.active
        
        # 设置表头格式
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
        
        # 设置数据格式
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(2, len(headers) + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    color = get_color_for_rank(value)
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # 调整列宽
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) * 1.2, 12)
        
        # 保存文件
        wb.save(RESULTS_EXCEL)
    
    logging.info(f"结果已更新到表格: {RESULTS_EXCEL}")

# 加载已有结果数据（用于断点续测）
existing_results = load_existing_results()

# 创建所有参数组合
combinations = []
for model_name, prompt_types in model_config.items():
    for prompt_type in prompt_types:
        for try_times in try_times_range:
            # 检查是否已有结果（断点续测）
            if model_name in existing_results and prompt_type in existing_results[model_name] and try_times in existing_results[model_name][prompt_type]:
                logging.info(f"跳过已有结果: model_name={model_name}, prompt_type={prompt_type}, try_times={try_times}, 结果={existing_results[model_name][prompt_type][try_times]}")
                continue
                
            combinations.append((model_name, prompt_type, try_times))

total_runs = len(combinations)

logging.info(f"开始批量运行，共{total_runs}个任务，使用{MAX_WORKERS}个并行线程")

# 定义CSV更新函数
def update_csv():
    with results_lock:
        df = pd.DataFrame(results)
        df.to_csv(os.path.join(run_dir, 'results.csv'), index=False)

# 定义执行任务的函数
def run_task(args):
    model_name, prompt_type, try_times = args
    
    # 检查是否已有结果（断点续测）
    if model_name in existing_results and prompt_type in existing_results[model_name] and try_times in existing_results[model_name][prompt_type]:
        rank = existing_results[model_name][prompt_type][try_times]
        logging.info(f"使用已有结果: model_name={model_name}, prompt_type={prompt_type}, try_times={try_times}, 结果={rank}")
        
        # 添加到结果中
        task_info = {
            'model_name': model_name,
            'prompt_type': prompt_type,
            'try_times': try_times,
            'rank': rank,
            'start_time': '',
            'end_time': '',
            'success': rank >= 0,
            'attempts': 0
        }
        
        with results_lock:
            results.append(task_info)
            
            # 更新结果表格
            if model_name not in existing_results:
                existing_results[model_name] = {}
            if prompt_type not in existing_results[model_name]:
                existing_results[model_name][prompt_type] = {}
            existing_results[model_name][prompt_type][try_times] = rank
            update_results_excel(existing_results)
            
        return task_info
    
    # 构建输出目录
    output_dir = f"result/shapez_2d/{model_name}/{prompt_type}"
    
    # 构建命令
    cmd = f"""python ladder_match.py \\
        --test_type ladder_match \\
        --source_data_path source_dataset/shapez_2d_source_data_min_1_max_100_each_100/source_data.jsonl \\
        --output_dir_path {output_dir} \\
        --service_name {SERVICE_NAME} \\
        --model_name {model_name} \\
        --response_name response \\
        --prompt_type {prompt_type} \\
        --min_steps 1 \\
        --max_steps 100 \\
        --each_number 100 \\
        --stop_if_accuracy_below_threshold True \\
        --accuracy_threshold 0.1 \\
        --max_waiting_time 200 \\
        --num_workers 5 \\
        --save_interval 1 \\
        --config_path reasoning_api_params \\
        --threshold 3 \\
        --initial_rank 1 \\
        --try_times {try_times}"""
    
    # 记录当前任务信息
    task_info = {
        'model_name': model_name,
        'prompt_type': prompt_type,
        'try_times': try_times,
        'start_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    # 创建任务日志文件
    log_file = os.path.join(run_dir, f"{model_name}_{prompt_type}_try{try_times}.log")
    
    # 定义结果文件路径
    result_file = os.path.join(output_dir, f'result_{model_name}_{prompt_type}_try{try_times}.json')
    
    # 执行命令，最多重试MAX_RETRIES次
    success = False
    attempt = 0
    rank = -1  # 默认为-1，表示出错
    
    while not success and attempt < MAX_RETRIES:
        attempt += 1
        logging.info(f"运行任务: model_name={model_name}, prompt_type={prompt_type}, try_times={try_times}, 尝试 {attempt}/{MAX_RETRIES}")
        
        try:
            # 执行命令并捕获输出
            with open(log_file, 'a') as f:
                f.write(f"=== 尝试 {attempt}/{MAX_RETRIES} at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n")
                f.write(f"命令: {cmd}\n\n")
                
                start_time = time.time()
                process = subprocess.Popen(
                    cmd, 
                    shell=True, 
                    stdout=subprocess.PIPE, 
                    stderr=subprocess.PIPE,
                    universal_newlines=True
                )
                
                for line in process.stdout:
                    f.write(line)
                    f.flush()
                
                for line in process.stderr:
                    f.write(f"ERROR: {line}")
                    f.flush()
                
                process.wait()
                end_time = time.time()
                
                f.write(f"\n命令执行时间: {end_time - start_time:.2f} 秒\n")
                f.write(f"返回码: {process.returncode}\n\n")
            
            # 不再依赖返回码，而是检查结果文件
            if os.path.exists(result_file):
                with open(result_file, 'r') as f:
                    try:
                        result_data = json.load(f)
                        rank = result_data.get('final_rank', -1)
                        error_info = result_data.get('error_info')
                        
                        if rank >= 0:  # 结果有效
                            success = True
                            logging.info(f"任务成功完成: model_name={model_name}, prompt_type={prompt_type}, try_times={try_times}, 达到层数: {rank}")
                        else:
                            logging.warning(f"任务失败: model_name={model_name}, prompt_type={prompt_type}, try_times={try_times}, 错误: {error_info}")
                            time.sleep(1)  # 等待1秒后重试
                    except json.JSONDecodeError:
                        logging.error(f"结果文件格式错误: {result_file}")
                        time.sleep(1)  # 等待1秒后重试
            else:
                logging.warning(f"未找到结果文件: {result_file}")
                time.sleep(1)  # 等待1秒后重试
                
        except Exception as e:
            logging.error(f"执行出错: {str(e)}")
            with open(log_file, 'a') as f:
                f.write(f"执行出错: {str(e)}\n")
            time.sleep(1)  # 等待1秒后重试
    
    # 记录执行结果
    task_info['end_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    task_info['success'] = success
    task_info['attempts'] = attempt
    task_info['log_file'] = log_file
    task_info['rank'] = rank
    
    # 线程安全地添加结果
    with results_lock:
        results.append(task_info)
        
        # 更新结果表格
        if model_name not in existing_results:
            existing_results[model_name] = {}
        if prompt_type not in existing_results[model_name]:
            existing_results[model_name][prompt_type] = {}
        existing_results[model_name][prompt_type][try_times] = rank
        update_results_excel(existing_results)
    
    # 更新CSV文件
    update_csv()
    
    return task_info

# 使用线程池执行任务
with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
    # 提交所有任务
    future_to_task = {executor.submit(run_task, args): args for args in combinations}
    
    # 使用tqdm显示进度
    with tqdm(total=total_runs, desc="正在运行任务") as pbar:
        for future in concurrent.futures.as_completed(future_to_task):
            args = future_to_task[future]
            try:
                task_info = future.result()
                pbar.update(1)
                model_name, prompt_type, try_times = args
                rank = task_info.get('rank', -1)
                pbar.set_description(f"完成: {model_name}_{prompt_type}_try{try_times} 结果: {rank}")
            except Exception as e:
                model_name, prompt_type, try_times = args
                logging.error(f"任务异常: model_name={model_name}, prompt_type={prompt_type}, try_times={try_times}, 错误: {str(e)}")
                pbar.update(1)

# 完成所有任务
logging.info(f"所有任务已完成，共{len(results)}个任务，成功{sum(1 for r in results if r.get('success', False))}个，失败{sum(1 for r in results if not r.get('success', False))}个")

# 生成摘要报告
success_rate = sum(1 for r in results if r.get('success', False)) / len(results) * 100 if results else 0
avg_attempts = sum(r.get('attempts', 0) for r in results) / len(results) if results else 0
avg_rank = sum(r.get('rank', 0) for r in results if r.get('rank', -1) >= 0) / sum(1 for r in results if r.get('rank', -1) >= 0) if sum(1 for r in results if r.get('rank', -1) >= 0) > 0 else 0

with open(os.path.join(run_dir, 'summary.txt'), 'w') as f:
    f.write(f"批量运行摘要报告\n")
    f.write(f"=================\n\n")
    f.write(f"运行时间: {run_id}\n")
    f.write(f"总任务数: {len(results)}\n")
    f.write(f"成功任务数: {sum(1 for r in results if r.get('success', False))}\n")
    f.write(f"失败任务数: {sum(1 for r in results if not r.get('success', False))}\n")
    f.write(f"成功率: {success_rate:.2f}%\n")
    f.write(f"平均尝试次数: {avg_attempts:.2f}\n")
    f.write(f"平均达到层数: {avg_rank:.2f}\n\n")
    
    if not all(r.get('success', False) for r in results):
        f.write("失败任务列表:\n")
        for r in results:
            if not r.get('success', False):
                f.write(f"- model_name={r['model_name']}, prompt_type={r['prompt_type']}, try_times={r['try_times']}, 结果={r.get('rank', -1)}\n")

print(f"批量运行已完成。详细日志请查看: {run_dir}")
print(f"结果表格已保存到: {RESULTS_EXCEL}")